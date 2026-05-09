#!/usr/bin/env python3
"""Prepare normalized historical RoboMaster CSV tables from official files."""

from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Iterable

from bs4 import BeautifulSoup
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"

RANK_ORDER = {
    "冠军": 1,
    "亚军": 2,
    "季军": 3,
    "殿军": 4,
    "八强": 8,
    "十六强": 16,
    "三十二强": 32,
}


def clean(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def split_school_team(value: str) -> tuple[str, str]:
    parts = [clean(part) for part in str(value).splitlines() if clean(part)]
    if len(parts) >= 2:
        return parts[0], parts[1]
    text = clean(value)
    return text, ""


def merged_span_map(ws) -> dict[tuple[int, int], int]:
    spans: dict[tuple[int, int], int] = {}
    for merged in ws.merged_cells.ranges:
        spans[(merged.min_row, merged.min_col)] = merged.max_row - merged.min_row + 1
    return spans


def result_row(
    *,
    year: str,
    stage: str,
    region: str,
    finish_group: str,
    school: str,
    team: str,
    award: str,
    qualification: str,
    source_id: str,
    notes: str = "",
) -> dict[str, str]:
    return {
        "year": year,
        "competition": "RMUC",
        "stage": stage,
        "region": region,
        "finish_group": finish_group,
        "finish_order": str(RANK_ORDER.get(finish_group, "")),
        "school": clean(school),
        "team": clean(team),
        "award": clean(award),
        "qualification": clean(qualification),
        "source_id": source_id,
        "notes": notes,
    }


def parse_2023_national() -> list[dict[str, str]]:
    path = DATA_DIR / "rmuc_2023_national_awards.xlsx"
    ws = load_workbook(path, data_only=True).active
    spans = merged_span_map(ws)
    rows: list[dict[str, str]] = []
    current_group = ""
    remaining = 0

    for row_idx in range(3, ws.max_row + 1):
        rank = clean(ws.cell(row_idx, 1).value)
        team_cell = ws.cell(row_idx, 2).value
        award = clean(ws.cell(row_idx, 3).value)
        if not team_cell or not award:
            continue
        if rank in RANK_ORDER:
            current_group = rank
            remaining = spans.get((row_idx, 1), 1) - 1
        elif remaining > 0:
            rank = current_group
            remaining -= 1
        else:
            rank = ""
        school, team = split_school_team(str(team_cell))
        note = "grouped_no_internal_rank" if rank in {"八强", "十六强", "三十二强"} else ""
        rows.append(
            result_row(
                year="2023",
                stage="全国赛",
                region="全国",
                finish_group=rank,
                school=school,
                team=team,
                award=award,
                qualification="",
                source_id="rmuc_2023_national_awards_xlsx",
                notes=note,
            )
        )
    return rows


def parse_2023_regional() -> list[dict[str, str]]:
    path = DATA_DIR / "rmuc_2023_regional_awards.xlsx"
    ws = load_workbook(path, data_only=True).active
    spans = merged_span_map(ws)
    blocks = [("南部赛区", 1), ("中部赛区", 6), ("北部赛区", 11)]
    rows: list[dict[str, str]] = []

    for region, col in blocks:
        current_group = ""
        remaining = 0
        for row_idx in range(3, ws.max_row + 1):
            rank = clean(ws.cell(row_idx, col).value)
            team_cell = ws.cell(row_idx, col + 1).value
            award = clean(ws.cell(row_idx, col + 2).value)
            qualification = clean(ws.cell(row_idx, col + 3).value)
            if not team_cell or not award or str(team_cell).startswith("*"):
                continue
            if rank in RANK_ORDER:
                current_group = rank
                remaining = spans.get((row_idx, col), 1) - 1
            elif remaining > 0:
                rank = current_group
                remaining -= 1
            else:
                rank = ""
            school, team = split_school_team(str(team_cell))
            note = "grouped_no_internal_rank" if rank in {"八强", "十六强"} else ""
            rows.append(
                result_row(
                    year="2023",
                    stage="区域赛",
                    region=region,
                    finish_group=rank,
                    school=school,
                    team=team,
                    award=award,
                    qualification=qualification,
                    source_id="rmuc_2023_regional_awards_xlsx",
                    notes=note,
                )
            )

    for row_idx in range(3, ws.max_row + 1):
        team_cell = ws.cell(row_idx, 16).value
        award = clean(ws.cell(row_idx, 17).value)
        if not team_cell or not award or str(team_cell).startswith("*"):
            continue
        school, team = split_school_team(str(team_cell))
        rows.append(
            result_row(
                year="2023",
                stage="邀请赛",
                region="邀请赛",
                finish_group="",
                school=school,
                team=team,
                award=award,
                qualification="",
                source_id="rmuc_2023_regional_awards_xlsx",
                notes="invitational_no_internal_rank",
            )
        )
    return rows


def table_texts(row) -> tuple[list[str], list[int]]:
    cells = row.find_all(["th", "td"])
    return [clean(cell.get_text(" ")) for cell in cells], [
        int(cell.get("rowspan", "1")) for cell in cells
    ]


def extract_region(text: str) -> str:
    match = re.search(r"(东部|中部|南部|北部)赛区", text)
    if match:
        return f"{match.group(1)}赛区"
    return clean(text)


def parse_awards_html(
    *,
    path: Path,
    year: str,
    stage: str,
    source_id: str,
    default_region: str,
    regional: bool,
) -> list[dict[str, str]]:
    soup = BeautifulSoup(path.read_text(encoding="utf-8"), "html.parser")
    table = soup.find("table")
    if table is None:
        return []

    rows: list[dict[str, str]] = []
    current_region = default_region
    current_group = ""
    remaining = 0

    for tr in table.find_all("tr"):
        cells, rowspans = table_texts(tr)
        if not cells:
            continue
        if len(cells) == 1 and "赛区" in cells[0] and "获奖名单" in cells[0]:
            current_region = extract_region(cells[0])
            current_group = ""
            remaining = 0
            continue
        if cells[0] in {"排名", "学校", "学校名称", ""} and len(cells) <= 2:
            continue
        if cells[0] in RANK_ORDER:
            group = cells[0]
            remaining = rowspans[0] - 1
            data = cells[1:]
            current_group = group
        elif remaining > 0:
            group = current_group
            remaining -= 1
            data = cells
        else:
            group = ""
            data = cells[1:] if cells[0] == "" else cells

        if regional:
            if len(data) < 3 or data[0] in {"排名", "学校"}:
                continue
            school, team, award = data[:3]
            qualification = data[3] if len(data) > 3 else ""
        else:
            if len(data) < 3 or data[0] in {"排名", "学校名称", "学校"}:
                continue
            school, team, award = data[:3]
            qualification = ""

        if not school or not team or not award or "联系我们" in school:
            continue
        note = "grouped_no_internal_rank" if group in {"八强", "十六强", "三十二强"} else ""
        rows.append(
            result_row(
                year=year,
                stage=stage,
                region=current_region,
                finish_group=group,
                school=school,
                team=team,
                award=award,
                qualification=qualification,
                source_id=source_id,
                notes=note,
            )
        )
    return rows


def parse_2023_complete_form() -> list[dict[str, str]]:
    path = DATA_DIR / "raw_rmuc_2023_complete_form_official.html"
    soup = BeautifulSoup(path.read_text(encoding="utf-8"), "html.parser")
    table = soup.find("table")
    rows: list[dict[str, str]] = []
    if table is None:
        return rows

    rank = 0
    for tr in table.find_all("tr")[1:]:
        cells, _ = table_texts(tr)
        if len(cells) < 6:
            continue
        rank += 1
        rows.append(
            {
                "year": "2023",
                "rank": str(rank),
                "school": cells[0],
                "team": cells[1],
                "metric": "complete_form_score",
                "score": cells[2],
                "project_doc_bonus": cells[3],
                "technical_solution_bonus": cells[4],
                "total_initial_gold_bonus": cells[5],
                "source_id": "rmuc_2023_complete_form",
                "notes": "",
            }
        )
    return rows


def parse_complete_form_scores_html(
    path: Path, year: str, source_id: str
) -> list[dict[str, str]]:
    soup = BeautifulSoup(path.read_text(encoding="utf-8"), "html.parser")
    table = soup.find("table")
    rows: list[dict[str, str]] = []
    if table is None:
        return rows

    rank = 0
    for tr in table.find_all("tr")[1:]:
        cells, _ = table_texts(tr)
        if len(cells) < 6:
            continue
        rank += 1
        rows.append(
            {
                "year": year,
                "rank": str(rank),
                "school": cells[0],
                "team": cells[1],
                "metric": "complete_form_score",
                "score": cells[2],
                "project_doc_bonus": cells[3],
                "technical_solution_bonus": cells[4],
                "total_initial_gold_bonus": cells[5],
                "source_id": source_id,
                "notes": "",
            }
        )
    return rows


def parse_2025_complete_form() -> list[dict[str, str]]:
    path = DATA_DIR / "raw_rmuc_2025_complete_form_official.html"
    soup = BeautifulSoup(path.read_text(encoding="utf-8"), "html.parser")
    tables = soup.find_all("table")
    if len(tables) < 2:
        return []

    pass_order: dict[tuple[str, str], str] = {}
    for tr in tables[0].find_all("tr")[1:]:
        cells, _ = table_texts(tr)
        if len(cells) >= 3:
            pass_order[(cells[1], cells[2])] = cells[0]

    rows: list[dict[str, str]] = []
    for tr in tables[1].find_all("tr")[1:]:
        cells, _ = table_texts(tr)
        if len(cells) < 7:
            continue
        school, team = cells[0], cells[1]
        rows.append(
            {
                "year": "2025",
                "rank": pass_order.get((school, team), ""),
                "school": school,
                "team": team,
                "metric": "complete_form_initial_gold_bonus",
                "score": cells[6],
                "project_doc_bonus": cells[3],
                "technical_solution_bonus": cells[5],
                "total_initial_gold_bonus": cells[6],
                "source_id": "rmuc_2025_complete_form",
                "notes": (
                    f"project_doc_grade={cells[2]};"
                    f"technical_solution_grade={cells[4]};"
                    "official_sequence_not_score_rank"
                ),
            }
        )
    return rows


def parse_2026_complete_form() -> list[dict[str, str]]:
    path = DATA_DIR / "complete_form_2026.csv"
    rows: list[dict[str, str]] = []
    with path.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            rows.append(
                {
                    "year": "2026",
                    "rank": row["complete_rank"],
                    "school": row["school"],
                    "team": row["team"],
                    "metric": "complete_form_rank",
                    "score": "",
                    "project_doc_bonus": "",
                    "technical_solution_bonus": "",
                    "total_initial_gold_bonus": "",
                    "source_id": "rmuc_2026_complete_form",
                    "notes": "rank_only_current_season",
                }
            )
    return rows


def parse_technical_awards_2024() -> list[dict[str, str]]:
    path = DATA_DIR / "raw_rmuc_2024_robot_competitive_awards_official.html"
    soup = BeautifulSoup(path.read_text(encoding="utf-8"), "html.parser")
    rows: list[dict[str, str]] = []
    for table in soup.find_all("table"):
        table_rows = table.find_all("tr")
        if not table_rows:
            continue
        category = clean(table_rows[0].get_text(" "))
        if not category or category in {"学校名称 队伍名称 奖项"}:
            continue
        for tr in table_rows[2:]:
            cells, _ = table_texts(tr)
            if len(cells) < 3:
                continue
            school, team, award = cells[:3]
            rows.append(
                {
                    "year": "2024",
                    "award_category": "机器人竞技奖",
                    "award_name": category,
                    "school": school,
                    "team": team,
                    "award": award,
                    "source_id": "rmuc_2024_robot_competitive_awards",
                    "notes": "",
                }
            )
    return rows


def static_technical_awards_2023() -> list[dict[str, str]]:
    raw = [
        ("战术", "最佳战术奖", "南方科技大学", "ARTINX"),
        ("技术报告", "最佳技术报告奖", "东北林业大学", "Ares"),
        ("技术报告", "最佳技术报告奖", "华中科技大学", "狼牙"),
        ("技术报告", "最佳技术报告奖", "南方科技大学", "ARTINX"),
        ("技术报告", "最佳技术报告奖", "太原科技大学", "NewMaker"),
        ("赛季规划", "最佳赛季规划奖", "西北工业大学", "WMJ"),
        ("赛季规划", "最佳赛季规划奖", "重庆大学", "MechaX"),
        ("赛季规划", "最佳赛季规划奖", "厦门大学", "RCS"),
        ("赛季规划", "最佳赛季规划奖", "辽宁科技大学", "COD"),
        ("赛季规划", "最佳赛季规划奖", "广东工业大学", "DynamicX"),
        ("赛季规划", "最佳赛季规划奖", "吉林大学", "TARS_Go"),
        ("赛季规划", "最佳赛季规划奖", "首都师范大学", "PIE"),
        ("设计创意", "最佳设计创意奖", "南京航空航天大学", "长空御风"),
        ("设计创意", "最佳设计创意奖", "深圳大学", "RobotPilots"),
        ("技术突破", "年度技术突破奖-二等奖-步兵机器人", "上海交通大学", "交龙"),
        ("技术突破", "年度技术突破奖-三等奖-英雄机器人", "华南农业大学", "Taurus"),
        ("技术突破", "年度技术突破奖-三等奖-飞镖系统", "华南理工大学", "华南虎"),
        ("技术突破", "年度技术突破奖-三等奖-飞镖系统", "南京航空航天大学", "长空御风"),
        ("技术突破", "年度技术突破奖-三等奖-工程机器人", "华南理工大学", "华南虎"),
        ("技术突破", "年度技术突破奖-三等奖-工程机器人", "南京航空航天大学", "长空御风"),
        ("技术突破", "年度技术突破奖-三等奖-工程机器人", "深圳大学", "RobotPilots"),
        ("技术突破", "年度技术突破奖-三等奖-哨兵机器人", "哈尔滨工业大学", "I Hiter"),
        ("技术突破", "年度技术突破奖-三等奖-哨兵机器人", "华南农业大学", "Taurus"),
        ("技术突破", "年度技术突破奖-三等奖-哨兵机器人", "深圳大学", "RobotPilots"),
    ]
    return [
        {
            "year": "2023",
            "award_category": category,
            "award_name": name,
            "school": school,
            "team": team,
            "award": "",
            "source_id": "rmuc_2023_comprehensive_awards",
            "notes": "",
        }
        for category, name, school, team in raw
    ]


def static_comprehensive_awards_2024_2025() -> list[dict[str, str]]:
    raw = [
        ("2024", "综合类-战术", "最佳战术奖", "北京理工大学", "追梦", "rmuc_2024_comprehensive_awards"),
        ("2024", "综合类-战术", "最佳战术奖", "中国石油大学（华东）", "RPS", "rmuc_2024_comprehensive_awards"),
        ("2024", "综合类-技术报告", "最佳技术报告奖", "太原理工大学", "TRoMaC", "rmuc_2024_comprehensive_awards"),
        ("2024", "综合类-技术报告", "最佳技术报告奖", "中国石油大学（华东）", "RPS", "rmuc_2024_comprehensive_awards"),
        ("2024", "综合类-技术报告", "最佳技术报告奖", "中南大学", "FYT", "rmuc_2024_comprehensive_awards"),
        ("2024", "综合类-年度技术突破", "二等奖 步兵机器人", "上海交通大学", "交龙", "rmuc_2024_comprehensive_awards"),
        ("2024", "综合类-年度技术突破", "二等奖 工程机器人", "中国石油大学（华东）", "RPS", "rmuc_2024_comprehensive_awards"),
        ("2024", "综合类-年度技术突破", "三等奖 工程机器人", "上海交通大学", "交龙", "rmuc_2024_comprehensive_awards"),
        ("2024", "综合类-年度技术突破", "三等奖 雷达", "哈尔滨工业大学（深圳）", "南工骁鹰", "rmuc_2024_comprehensive_awards"),
        ("2024", "综合类-年度技术突破", "三等奖 飞镖系统", "浙江大学", "Hello World", "rmuc_2024_comprehensive_awards"),
        ("2025", "综合类-战术", "最佳战术奖", "华南理工大学", "华南虎", "rmuc_2025_comprehensive_awards"),
        ("2025", "综合类-技术报告", "最佳技术报告奖 RMB5,000（税前）", "中国科学技术大学", "RoboWalker", "rmuc_2025_comprehensive_awards"),
        ("2025", "综合类-技术报告", "最佳技术报告奖 RMB3,000（税前）", "哈尔滨工业大学", "I Hiter", "rmuc_2025_comprehensive_awards"),
        ("2025", "综合类-技术报告", "最佳技术报告奖 RMB3,000（税前）", "西北工业大学", "WMJ", "rmuc_2025_comprehensive_awards"),
        ("2025", "综合类-技术报告", "最佳技术报告奖 RMB3,000（税前）", "中国石油大学（华东）", "RPS", "rmuc_2025_comprehensive_awards"),
        ("2025", "综合类-技术报告", "最佳技术报告奖 RMB2,000（税前）", "南昌大学", "Passion", "rmuc_2025_comprehensive_awards"),
        ("2025", "综合类-年度技术突破", "二等奖 飞镖系统", "南京航空航天大学", "长空御风", "rmuc_2025_comprehensive_awards"),
        ("2025", "综合类-年度技术突破", "二等奖 飞镖系统", "浙江大学", "Hello World", "rmuc_2025_comprehensive_awards"),
        ("2025", "综合类-年度技术突破", "三等奖 英雄机器人", "上海交通大学", "交龙", "rmuc_2025_comprehensive_awards"),
        ("2025", "综合类-年度技术突破", "三等奖 哨兵机器人", "中国科学技术大学", "RoboWalker", "rmuc_2025_comprehensive_awards"),
    ]
    return [
        {
            "year": year,
            "award_category": category,
            "award_name": name,
            "school": school,
            "team": team,
            "award": "",
            "source_id": source_id,
            "notes": "",
        }
        for year, category, name, school, team, source_id in raw
    ]


def static_prediction_feature_weights() -> list[dict[str, str]]:
    return [
        ("university_score_2025", "0.22", "higher_better", "RoboMaster高校积分榜是跨赛事历史实力先验"),
        ("complete_form_rank_2026", "0.16", "lower_better", "赛前完整形态越靠前，机器人完成度和稳定性先验越强"),
        ("initial_gold_bonus_2026", "0.12", "higher_better", "项目文档与技术方案影响每局初始金币"),
        ("national_top8_count_last_3y", "0.16", "higher_better", "全国赛八强频次代表强队稳定性"),
        ("regional_advancement_count_last_3y", "0.10", "higher_better", "区域赛晋级频次代表下限"),
        ("technical_award_count_last_3y", "0.08", "higher_better", "机器人竞技奖/技术奖代表专项能力"),
        ("recent_match_rating", "0.18", "higher_better", "逐场赛果补齐后替换为Glicko/Elo主特征"),
        ("upset_risk_index", "-0.10", "higher_risk", "高风险场降低确定性而非直接改变强弱方向"),
    ]


def static_backtest_plan() -> list[dict[str, str]]:
    return [
        ("split_2023_to_2024", "2023-01-01", "2023-12-31", "2024-01-01", "2024-12-31", "训练2023，测试2024"),
        ("split_2023_2024_to_2025", "2023-01-01", "2024-12-31", "2025-01-01", "2025-12-31", "训练2023-2024，测试2025"),
        ("rolling_2026_regional", "2023-01-01", "2026-04-30", "2026-05-01", "2026-06-30", "2026区域赛滚动预测"),
    ]


def static_match_data_sources() -> list[dict[str, str]]:
    return [
        ("official_live_data", "https://www.robomaster.com/live/data", "赛程赛果入口", "2023-2026", "4", "5", "历史结构化内容需要进一步抓取/接口解析", "robomaster_live_data"),
        ("official_bilibili_replays", "https://space.bilibili.com/20554233", "官方逐场回放标题可抽对阵", "2023-2025", "2", "4.5", "比分多数需OCR或交叉校验", "bilibili_rmuc_replays"),
        ("community_results_summary", "https://bbs.robomaster.com/article/25006", "非官方赛果汇总与腾讯文档", "2015-2025", "3", "3.5", "适合补winner/比分，必须保留置信度", "rm_community_results_2015_2025"),
    ]


def static_upset_features() -> list[dict[str, str]]:
    return [
        ("rating_uncertainty", "rating", "higher_means_more_upset_risk", "Glicko RD/TrueSkill sigma 高说明强弱判断不稳"),
        ("rating_gap_small", "matchup", "higher_means_more_upset_risk", "双方rating差距小，强弱标签本身脆弱"),
        ("technical_gap_small", "matchup", "higher_means_more_upset_risk", "完整形态/初始金币差距小，硬实力先验不足"),
        ("underdog_recent_momentum", "form", "higher_means_more_upset_risk", "弱队近期区域赛晋级或技术奖表现上升"),
        ("favorite_recent_decline", "form", "higher_means_more_upset_risk", "强队近一年名次下降或技术信号走弱"),
        ("low_match_count", "data_quality", "higher_means_more_upset_risk", "样本不足时不应高置信预测"),
        ("stage_variance", "context", "higher_means_more_upset_risk", "淘汰赛、BO短局、关键兵种故障会放大方差"),
    ]


def write_csv(path: Path, rows: Iterable[dict[str, str]], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    competition_rows: list[dict[str, str]] = []
    competition_rows.extend(parse_2023_national())
    competition_rows.extend(parse_2023_regional())
    competition_rows.extend(
        parse_awards_html(
            path=DATA_DIR / "raw_rmuc_2024_national_awards_official.html",
            year="2024",
            stage="全国赛",
            source_id="rmuc_2024_national_awards",
            default_region="全国",
            regional=False,
        )
    )
    competition_rows.extend(
        parse_awards_html(
            path=DATA_DIR / "raw_rmuc_2024_regional_awards_official.html",
            year="2024",
            stage="区域赛",
            source_id="rmuc_2024_regional_awards",
            default_region="东部赛区",
            regional=True,
        )
    )
    competition_rows.extend(
        parse_awards_html(
            path=DATA_DIR / "raw_rmuc_2025_national_awards_official.html",
            year="2025",
            stage="全国赛",
            source_id="rmuc_2025_national_awards",
            default_region="全国",
            regional=False,
        )
    )
    competition_rows.extend(
        parse_awards_html(
            path=DATA_DIR / "raw_rmuc_2025_regional_awards_official.html",
            year="2025",
            stage="区域赛",
            source_id="rmuc_2025_regional_awards",
            default_region="南部赛区",
            regional=True,
        )
    )
    write_csv(
        DATA_DIR / "historical_competition_results.csv",
        competition_rows,
        [
            "year",
            "competition",
            "stage",
            "region",
            "finish_group",
            "finish_order",
            "school",
            "team",
            "award",
            "qualification",
            "source_id",
            "notes",
        ],
    )

    technical_rows = (
        static_technical_awards_2023()
        + parse_technical_awards_2024()
        + static_comprehensive_awards_2024_2025()
    )
    write_csv(
        DATA_DIR / "historical_technical_awards.csv",
        technical_rows,
        ["year", "award_category", "award_name", "school", "team", "award", "source_id", "notes"],
    )

    preseason_rows = (
        parse_2023_complete_form()
        + parse_complete_form_scores_html(
            DATA_DIR / "raw_rmuc_2024_complete_form_scores_official.html",
            "2024",
            "rmuc_2024_complete_form_scores",
        )
        + parse_2025_complete_form()
        + parse_2026_complete_form()
    )
    write_csv(
        DATA_DIR / "historical_preseason_assessments.csv",
        preseason_rows,
        [
            "year",
            "rank",
            "school",
            "team",
            "metric",
            "score",
            "project_doc_bonus",
            "technical_solution_bonus",
            "total_initial_gold_bonus",
            "source_id",
            "notes",
        ],
    )

    write_csv(
        DATA_DIR / "prediction_feature_weights.csv",
        (
            {
                "feature_name": name,
                "baseline_weight": weight,
                "direction": direction,
                "rationale": rationale,
                "source_id": "model_design_internal",
            }
            for name, weight, direction, rationale in static_prediction_feature_weights()
        ),
        ["feature_name", "baseline_weight", "direction", "rationale", "source_id"],
    )

    write_csv(
        DATA_DIR / "model_backtest_plan.csv",
        (
            {
                "split_name": name,
                "train_start_date": train_start,
                "train_end_date": train_end,
                "test_start_date": test_start,
                "test_end_date": test_end,
                "target_metric": "match_winner",
                "primary_metrics": "accuracy,brier,log_loss,accuracy_at_confidence_buckets",
                "source_id": "model_design_internal",
                "notes": notes,
            }
            for name, train_start, train_end, test_start, test_end, notes in static_backtest_plan()
        ),
        [
            "split_name",
            "train_start_date",
            "train_end_date",
            "test_start_date",
            "test_end_date",
            "target_metric",
            "primary_metrics",
            "source_id",
            "notes",
        ],
    )

    write_csv(
        DATA_DIR / "match_data_sources.csv",
        (
            {
                "source_name": name,
                "url": url,
                "available_fields": fields,
                "coverage": coverage,
                "collection_difficulty_1_easy_5_hard": difficulty,
                "reliability_1_low_5_high": reliability,
                "notes": notes,
                "source_id": source_id,
            }
            for name, url, fields, coverage, difficulty, reliability, notes, source_id in static_match_data_sources()
        ),
        [
            "source_name",
            "url",
            "available_fields",
            "coverage",
            "collection_difficulty_1_easy_5_hard",
            "reliability_1_low_5_high",
            "notes",
            "source_id",
        ],
    )

    write_csv(
        DATA_DIR / "upset_model_features.csv",
        (
            {
                "feature_name": name,
                "feature_family": family,
                "signal_direction": direction,
                "description": description,
                "source_id": "model_design_internal",
            }
            for name, family, direction, description in static_upset_features()
        ),
        ["feature_name", "feature_family", "signal_direction", "description", "source_id"],
    )

    empty_tables = {
        "match_results.csv": [
            "match_id",
            "year",
            "stage",
            "region",
            "round",
            "match_no",
            "team_a",
            "team_b",
            "winner",
            "score_a",
            "score_b",
            "source_url",
            "confidence",
            "source_id",
            "notes",
        ],
        "team_strength_ratings.csv": [
            "rating_id",
            "rating_system",
            "rating_version",
            "school",
            "team",
            "as_of_date",
            "rating_mu",
            "rating_sigma",
            "rating_rd",
            "games_played",
            "source_id",
            "notes",
        ],
        "model_predictions.csv": [
            "prediction_id",
            "model_run_id",
            "match_id",
            "prediction_created_at",
            "data_cutoff",
            "team_a",
            "team_b",
            "p_team_a_win",
            "p_team_b_win",
            "predicted_winner",
            "confidence",
            "upset_risk_index",
            "actual_winner",
            "correct",
            "source_id",
            "notes",
        ],
        "model_backtests.csv": [
            "backtest_id",
            "model_run_id",
            "split_name",
            "n_matches",
            "coverage",
            "accuracy",
            "accuracy_top_confidence_20pct",
            "brier",
            "log_loss",
            "high_confidence_error_rate",
            "source_id",
            "notes",
        ],
    }
    for file_name, fieldnames in empty_tables.items():
        write_csv(DATA_DIR / file_name, [], fieldnames)

    print(f"historical_competition_results: {len(competition_rows)} rows")
    print(f"historical_technical_awards: {len(technical_rows)} rows")
    print(f"historical_preseason_assessments: {len(preseason_rows)} rows")


if __name__ == "__main__":
    main()
